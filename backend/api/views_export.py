import csv
import io
from django.http import HttpResponse, FileResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import logging

from .models import Quotation, DocumentHistory
from .utils.pdf_generator import generate_quotation_pdf

logger = logging.getLogger(__name__)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_quotation_pdf(request, pk):
    quotation = get_object_or_404(Quotation, pk=pk)
    user = request.user

    # Ownership check: only the creator, managers, or admins can export
    if quotation.createdBy != user and user.role not in ('Manager', 'Admin'):
        return Response({"detail": "Not authorized to export this quotation."}, status=status.HTTP_403_FORBIDDEN)

    try:
        pdf_buffer = generate_quotation_pdf(quotation)
    except Exception as e:
        logger.error(f"PDF generation failed for quotation {quotation.quotationNumber}: {e}")
        return Response({"detail": "PDF generation failed. Please try again."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    DocumentHistory.objects.create(
        quotation=quotation,
        user=request.user,
        action="PDF generated"
    )

    filename = f"Quotation_{quotation.quotationNumber}.pdf"
    response = FileResponse(pdf_buffer, as_attachment=True, filename=filename)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_excel(request):
    user = request.user

    # Employees only export their own; managers/admins export all
    if user.role in ('Manager', 'Admin'):
        quotations = Quotation.objects.select_related('clientInfo').order_by('-createdAt')
    else:
        quotations = Quotation.objects.select_related('clientInfo').filter(createdBy=user).order_by('-createdAt')

    wb = Workbook()
    ws = wb.active
    ws.title = "Quotations Export"

    headers = ['Quotation Number', 'Client Name', 'Issue Date', 'Expiry Date', 'Status', 'Subtotal', 'GST', 'Grand Total']
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="111111")
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill

    for q in quotations:
        try:
            client_name = q.clientInfo.companyName
        except Exception:
            client_name = 'N/A'

        # GST is calculated on after-discount amount
        discount_amount = float(q.subtotal) * (float(q.discountPercent) / 100)
        after_discount = float(q.subtotal) - discount_amount
        gst_amount = after_discount * (float(q.gstPercent) / 100)

        ws.append([
            q.quotationNumber,
            client_name,
            q.issueDate.strftime('%Y-%m-%d'),
            q.expiryDate.strftime('%Y-%m-%d'),
            q.status,
            float(q.subtotal),
            round(gst_amount, 2),
            float(q.grandTotal)
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Bulk log export action (single entry, not per-quotation)
    logger.info(f"Excel export of {quotations.count()} quotations by {user.username}")

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Quotations_Export.xlsx"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_csv(request):
    user = request.user

    if user.role in ('Manager', 'Admin'):
        quotations = Quotation.objects.select_related('clientInfo').order_by('-createdAt')
    else:
        quotations = Quotation.objects.select_related('clientInfo').filter(createdBy=user).order_by('-createdAt')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="Quotations_Export.csv"'

    writer = csv.writer(response)
    writer.writerow(['Quotation Number', 'Client Name', 'Issue Date', 'Expiry Date', 'Status', 'Subtotal', 'GST', 'Grand Total'])

    for q in quotations:
        try:
            client_name = q.clientInfo.companyName
        except Exception:
            client_name = 'N/A'

        # Correct GST calculation (on after-discount amount)
        discount_amount = float(q.subtotal) * (float(q.discountPercent) / 100)
        after_discount = float(q.subtotal) - discount_amount
        gst_amount = after_discount * (float(q.gstPercent) / 100)

        writer.writerow([
            q.quotationNumber,
            client_name,
            q.issueDate.strftime('%Y-%m-%d'),
            q.expiryDate.strftime('%Y-%m-%d'),
            q.status,
            f"{q.subtotal:.2f}",
            f"{gst_amount:.2f}",
            f"{q.grandTotal:.2f}"
        ])

    logger.info(f"CSV export of {quotations.count()} quotations by {user.username}")
    return response
